import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from scipy.interpolate import interp1d
from openpyxl import load_workbook
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import pagesizes
from reportlab.lib.units import inch
import os
import cma
import grcwa
from config import *
from geometry_functions import *
from geometry_visuals import *
from rcwa_machinery import *
from plot_spectra import *
from sweep_functions import *



lambda0 = 1.5

lambdas = np.linspace(1.4, 1.6, 2000) 


hs_dbr = lambda0/(4*3.4778)
hs_SiO2_dbr = lambda0/(4*1.45)

geometry_func = get_epgrid_3x3 






'''Geometry 1'''

geometry_params1 = [0.29223819, 0.96316646, 0.01451838, 0.51136224, 0.48586453, 0.94619988,
 0.84388626, 0.68818074, 0.9603793]
normal_params1 = [0.95000057, 0.95000057, 0.23473175, hs_dbr, hs_SiO2_dbr] #L1,L2,h,h_Sidbr,h_Sio2




'''Gives you phase as a function of wavelengths'''
phis1,r_amp1 = compute_phase(geometry_func, geometry_params1, lambdas, normal_params1, DBR_PAIRS)














# ==============================
#  ULTRAFAST PULSE PROPAGATION
# ==============================

c = 3e8

# --------------------------------------------------
# 1) Convert wavelength (micron) -> angular frequency
# --------------------------------------------------

lambdas_m = lambdas * 1e-6
omega = 2*np.pi*c / lambdas_m

# 1. Update your lambdas at the top of your script
 # Broad range for smooth dt ~ 10-15fs

# ... (compute phis as normal for your geometry) ...

# --------------------------------------------------
# 2) Fixed Interpolation logic
# --------------------------------------------------

# Important: Sort omega and phis before interpolating
sort_idx = np.argsort(omega)
omega_sorted = omega[sort_idx]
phis1_sorted = phis1[sort_idx]


# Define uniform grid over the full 1.4-1.6 range
Nw = 8192
omega_uniform = np.linspace(omega_sorted.min(), omega_sorted.max(), Nw)

# FIX: Use fill_value=(left, right) to pad with the first/last real phase values
# instead of "extrapolate" which creates fake linear slopes.
interp_phi1 = interp1d(omega_sorted, phis1_sorted, kind='cubic', 
                      bounds_error=False, fill_value=(phis1_sorted[0], phis1_sorted[-1]))



phi1_uniform = interp_phi1(omega_uniform)


# --------------------------------------------------
# 3) Verification: Zeroing out phase (Optional but Safer)
# --------------------------------------------------
# This ensures that wavelengths far from 1.51um don't accidentally contribute
# to a shift, though the Gaussian Ein already handles this.
mask = (omega_uniform < omega_sorted.min()) | (omega_uniform > omega_sorted.max())
phi1_uniform[mask] = 0



# --------------------------------------------------
# 3) Define Gaussian spectral pulse
# --------------------------------------------------

lambda_center = 1.499e-6
omega0 = 2*np.pi*c / lambda_center

# 10 nm FWHM bandwidth
delta_lambda = 6e-9
delta_omega = (2*np.pi*c/(lambda_center**2)) * delta_lambda

sigma_omega = delta_omega / (2*np.sqrt(2*np.log(2)))

Ein = np.exp(-(omega_uniform - omega0)**2 / (2*sigma_omega**2))

# --------------------------------------------------
# 4) Apply phase
# --------------------------------------------------

Eout1 = Ein * np.exp(1j*phi1_uniform)


# --------------------------------------------------
# 5) Inverse Fourier Transform
# --------------------------------------------------

import scipy.signal as signal

# ... (keep your steps 1-4) ...

# --------------------------------------------------
# CLEAN FREQUENCY-DOMAIN → TIME-DOMAIN TRANSFORM
# --------------------------------------------------

# Convert angular frequency to ordinary frequency (Hz)
freq = omega_uniform / (2*np.pi)

# Frequency spacing
df = freq[1] - freq[0]

# Time axis (consistent with freq spacing)
t = np.fft.fftshift(np.fft.fftfreq(Nw, d=df))

# --- Properly scaled IFFT ---
# Multiply by df so energy is preserved in continuous sense

Ein_t = np.fft.fftshift(
    np.fft.ifft(np.fft.ifftshift(Ein))
) * Nw * df


Eout1_t = np.fft.fftshift(
    np.fft.ifft(np.fft.ifftshift(Eout1))
) * Nw * df



# Intensities
Iin = np.abs(Ein_t)**2
Iout1 = np.abs(Eout1_t)**2


Ein_energy = np.trapz(Iin, t)





# --------------------------------------------------
# 7) Extract Peak Times (Group Delay from Pulse Peak)
# --------------------------------------------------

idx_in  = np.argmax(Iin)
idx_1   = np.argmax(Iout1)


t_in_peak = t[idx_in]
t_1_peak  = t[idx_1]


# Delays in femtoseconds
delay1_fs = (t_1_peak - t_in_peak) * 1e15


print(f"Geometry 3 delay: {delay1_fs:.2f} fs")






# --------------------------------------------------
# 8) NEW: Plot Time Domain Pulses (ZOOMED IN FS)
# --------------------------------------------------
scale = np.max(Iin)

plt.figure(figsize=(10,5))
plt.plot(t*1e15, Iin/scale, 'k--', label="Input")
plt.plot(t*1e15, Iout1/scale, label="Geometry 3 (cubic)")

plt.xlim(-2000,2000)
plt.xlabel("Time (fs)")
plt.ylabel("Intensity (normalized to input peak)")
plt.legend()
plt.show()












# --------------------------------------------------
# 12) Group Delay Difference Calculation
# --------------------------------------------------
# Find the peak position of each pulse in time
# Add this inside Step 12 for high-precision delay calculation
t_com_in = np.sum(t * Iin) / np.sum(Iin)
###t_com_out1 = np.sum(t * Iout1) / np.sum(Iout1)








# ==========================================================
# Compute Group Delay
# ==========================================================

dphi1 = np.gradient(phi1_uniform, omega_uniform)


# Convert to femtoseconds
tau1_fs = dphi1 * 1e15


# Central frequency index
idx0 = np.argmin(np.abs(omega_uniform - omega0))

print("Group delay at center (Geometry 1):", tau1_fs[idx0], "fs")


# ==========================================================
# Extract Pulse Peak Delays (Already Done Before)
# ==========================================================

idx_in  = np.argmax(Iin)
idx_1   = np.argmax(Iout1)


t_in_peak = t[idx_in]
t_1_peak  = t[idx_1]


delay1_fs = (t_1_peak - t_in_peak) * 1e15


# ==========================================================
# Publication Quality Multi-Panel Figure
# ==========================================================


# Convert wavelength window to angular frequency
c = 3e8

lambda_min = 1.493e-6
lambda_max = 1.505e-6

omega_min = 2*np.pi*c / lambda_max  # careful: inverse relation
omega_max = 2*np.pi*c / lambda_min

plt.rcParams.update({
    "font.size": 12,
    "axes.labelsize": 13,
    "legend.fontsize": 11,
    "xtick.labelsize": 11,
    "ytick.labelsize": 11,
    "axes.linewidth": 1.2
})

fig, ax = plt.subplots(figsize=(6, 4))

ax.plot(omega_uniform, phi1_uniform, linewidth=2, label="Geometry 3")


# Shade 1.50–1.52 µm region
ax.axvspan(omega_min, omega_max, color='grey', alpha=0.15,
           label="1.493–1.505 µm")

ax.set_ylabel("Phase φ(ω) (rad)")
ax.set_xlabel("Angular Frequency ω (rad/s)")
ax.set_title("Spectral Phase")
ax.legend(frameon=False)

plt.tight_layout()
plt.show()


# ==========================================================
# Restrict Analysis to Linear Phase Region (1.50–1.52 µm)
# ==========================================================

# Convert omega back to wavelength (microns)
lambda_uniform = 2*np.pi*c / omega_uniform * 1e6

# Define cubic window
lambda_min = 1.493
lambda_max = 1.505

mask_linear = (lambda_uniform >= lambda_min) & (lambda_uniform <= lambda_max)

# Extract region
omega_lin = omega_uniform[mask_linear]
phi1_lin  = phi1_uniform[mask_linear]


# ==========================================================
# Linear Fit to Phase  φ = aω + b
# ==========================================================

coef1 = np.polyfit(omega_lin, phi1_lin, 1)


slope1 = coef1[0]


# Group delay = dφ/dω = slope
tau1_linear_fs = slope1 * 1e15


print("Linear-fit Group Delay (Geometry 1):", tau1_linear_fs, "fs")




# ==========================================================
# FUNCTION: Effective Group Delay + GDD
# ==========================================================

def compute_dispersion_metrics(phi, omega, Ein):

    # First derivative (group delay)
    dphi = np.gradient(phi, omega)
    tau = dphi  # seconds

    # Second derivative (GDD)
    d2phi = np.gradient(dphi, omega)
    GDD = d2phi  # seconds^2

    # Spectral weight
    S = np.abs(Ein)**2

    # Weighted averages
    tau_eff = np.trapz(tau * S, omega) / np.trapz(S, omega)
    GDD_eff = np.trapz(GDD * S, omega) / np.trapz(S, omega)

    return tau_eff, GDD_eff





tau1_eff, GDD1_eff = compute_dispersion_metrics(phi1_uniform, omega_uniform, Ein)


tau1_eff_fs = tau1_eff * 1e15


GDD1_eff_fs2 = GDD1_eff * (1e15)**2


print("Effective GD G1 (fs):", tau1_eff_fs)

print("Effective GDD G1 (fs^2):", GDD1_eff_fs2)








# ==========================================================
# Compute Natural Group Delay and GDD
# ==========================================================

dphi1 = np.gradient(phi1_uniform, omega_uniform)


d2phi1 = np.gradient(dphi1, omega_uniform)


tau1_fs_full = dphi1 * 1e15


GDD1_fs2_full = d2phi1 * (1e15)**2


# Restrict to linear window
tau1_fs_lin = tau1_fs_full[mask_linear]


GDD1_fs2_lin = GDD1_fs2_full[mask_linear]



# ==========================================================
# Publication Multi-Panel (Phase + GD + GDD)
# ==========================================================

plt.rcParams.update({
    "font.size": 12,
    "axes.labelsize": 13,
    "legend.fontsize": 11,
    "xtick.labelsize": 11,
    "ytick.labelsize": 11,
    "axes.linewidth": 1.2
})

fig, axs = plt.subplots(3, 1, figsize=(6, 9))

# ----------------------------------------------------------
# (a) Spectral Phase
# ----------------------------------------------------------
axs[0].plot(omega_lin, phi1_lin, linewidth=2, label="Geometry 3")





axs[0].set_ylabel("Phase φ(ω) (rad)")
axs[0].set_title("(a) Spectral Phase (1.493–1.505 µm)")
axs[0].legend(frameon=False)


# ----------------------------------------------------------
# (b) Group Delay
# ----------------------------------------------------------
axs[1].plot(omega_lin, tau1_fs_lin, linewidth=2, label="Geometry 3")


axs[1].axhline(tau1_eff_fs, linestyle='--', linewidth=1,
               label=f"Effective GD G3 = {tau1_eff_fs:.1f} fs")



axs[1].set_ylabel("Group Delay (fs)")
axs[1].set_title("(b) Group Delay dφ/dω")
axs[1].legend(frameon=False)


# ----------------------------------------------------------
# (c) GDD
# ----------------------------------------------------------
axs[2].plot(omega_lin, GDD1_fs2_lin, linewidth=2, label="Geometry 3")


axs[2].set_ylabel("GDD (fs²)")
axs[2].set_xlabel("Angular Frequency ω (rad/s)")
axs[2].set_title("(c) Group Delay Dispersion d²φ/dω²")
axs[2].legend(frameon=False)

plt.tight_layout()
plt.show()









# --------------------------------------------------
# Function to compute FWHM
# --------------------------------------------------

def compute_fwhm(t, I):
    half_max = np.max(I) / 2
    indices = np.where(I >= half_max)[0]
    if len(indices) < 2:
        return 0
    return t[indices[-1]] - t[indices[0]]





fwhm_in  = compute_fwhm(t, Iin)
fwhm_1   = compute_fwhm(t, Iout1)


print("FWHM_in  (fs):", fwhm_in*1e15)
print("FWHM_out1(fs):", fwhm_1*1e15)


print("Broadening factor G1:", fwhm_1 / fwhm_in)







def compute_centroid(t, I):
    return np.trapz(t * I, t) / np.trapz(I, t)

centroid_in  = compute_centroid(t, Iin)
centroid_1   = compute_centroid(t, Iout1)


print("Centroid delay G1 (fs):", (centroid_1 - centroid_in)*1e15)




def compute_rms_width(t, I):
    centroid = compute_centroid(t, I)
    variance = np.trapz((t - centroid)**2 * I, t) / np.trapz(I, t)
    return np.sqrt(variance)

rms_in  = compute_rms_width(t, Iin)
rms_1   = compute_rms_width(t, Iout1)


print("RMS width in (fs):", rms_in*1e15)
print("RMS width G1 (fs):", rms_1*1e15)





def compute_skewness(t, I):
    centroid = compute_centroid(t, I)
    rms = compute_rms_width(t, I)
    third_moment = np.trapz((t - centroid)**3 * I, t) / np.trapz(I, t)
    return third_moment / rms**3

skew_in  = compute_skewness(t, Iin)
skew_1   = compute_skewness(t, Iout1)


print("Skewness in :", skew_in)
print("Skewness G1:", skew_1)








# ----------------------------------------------------------
# 1) GROUP DELAY from spectral phase
# ----------------------------------------------------------

dphi = np.gradient(phi1_uniform, omega_uniform)
tau_fs = dphi * 1e15

idx0 = np.argmin(np.abs(omega_uniform - omega0))
group_delay_fs = tau_fs[idx0]

# ----------------------------------------------------------
# 2) Centroid delay
# ----------------------------------------------------------

def compute_centroid(t, I):
    return np.trapz(t * I, t) / np.trapz(I, t)

centroid_in  = compute_centroid(t, Iin)
centroid_out = compute_centroid(t, Iout1)

centroid_delay_fs = (centroid_out - centroid_in) * 1e15

# ----------------------------------------------------------
# 3) FWHM
# ----------------------------------------------------------

def compute_fwhm(t, I):
    half_max = np.max(I) / 2
    indices = np.where(I >= half_max)[0]
    if len(indices) < 2:
        return 0
    return t[indices[-1]] - t[indices[0]]

fwhm_in  = compute_fwhm(t, Iin)
fwhm_out = compute_fwhm(t, Iout1)

fwhm_in_fs  = fwhm_in * 1e15
fwhm_out_fs = fwhm_out * 1e15

broadening_factor = fwhm_out / fwhm_in if fwhm_in != 0 else 0

# ----------------------------------------------------------
# 4) RMS width
# ----------------------------------------------------------

def compute_rms_width(t, I):
    centroid = compute_centroid(t, I)
    variance = np.trapz((t - centroid)**2 * I, t) / np.trapz(I, t)
    return np.sqrt(variance)

rms_in  = compute_rms_width(t, Iin)
rms_out = compute_rms_width(t, Iout1)

rms_in_fs  = rms_in * 1e15
rms_out_fs = rms_out * 1e15

# ----------------------------------------------------------
# 5) Skewness
# ----------------------------------------------------------

def compute_skewness(t, I):
    centroid = compute_centroid(t, I)
    rms = compute_rms_width(t, I)
    third_moment = np.trapz((t - centroid)**3 * I, t) / np.trapz(I, t)
    return third_moment / rms**3 if rms != 0 else 0

skewness = compute_skewness(t, Iout1)

# ----------------------------------------------------------
# 6) Define Metadata (EDIT THESE EACH RUN)
# ----------------------------------------------------------

phase_type = "Cubic"            # Linear / Quadratic / Cubic
geometry_type = "Patterned Square"
geometry_id = "G3"
pulse_bandwidth_nm = 6        # Your bandwidth in nm

# ----------------------------------------------------------
# 7) Append to Existing Excel File
# ----------------------------------------------------------

file_path = r"D:\msc\Research\presentations\week17\comparison\Pulse.xlsx"

if not os.path.exists(file_path):
    raise FileNotFoundError("Pulse.xlsx not found at specified path.")

wb = load_workbook(file_path)
ws = wb.active

new_row = [
    phase_type,
    geometry_type,
    geometry_id,
    pulse_bandwidth_nm,
    tau1_eff_fs,          # Effective GD
    centroid_delay_fs,    # Time-domain check
    fwhm_in_fs,
    fwhm_out_fs,
    broadening_factor,
    rms_in_fs,
    rms_out_fs,
    skewness,
    GDD1_eff_fs2          # NEW COLUMN
]

ws.append(new_row)
wb.save(file_path)

print("Data successfully appended to Pulse.xlsx")







# ==========================================================
# Compute All Pulse Metrics
# ==========================================================





# ==========================================================
# JOURNAL-STYLE TABLE GENERATOR FROM Pulse.xlsx
# ==========================================================




# ----------------------------------------------------------
# EDIT THESE PATHS
# ----------------------------------------------------------

excel_path = r"D:\msc\Research\presentations\week17\comparison\Pulse.xlsx"
pdf_path   = r"D:\msc\Research\presentations\week17\comparison\Pulse_Publication_Table.pdf"

if not os.path.exists(excel_path):
    raise FileNotFoundError("Pulse.xlsx not found at specified path.")

# ----------------------------------------------------------
# Load Excel Data
# ----------------------------------------------------------

wb = load_workbook(excel_path)
ws = wb.active

data = []

for row in ws.iter_rows(values_only=True):
    formatted_row = []
    for cell in row:
        if isinstance(cell, float):
            # Clean scientific formatting
            if abs(cell) < 1e-3 or abs(cell) > 1e4:
                formatted_row.append(f"{cell:.3e}")
            else:
                formatted_row.append(f"{cell:.3f}")
        else:
            formatted_row.append(str(cell))
    data.append(formatted_row)

# ----------------------------------------------------------
# Create PDF Document
# ----------------------------------------------------------

doc = SimpleDocTemplate(
    pdf_path,
    pagesize=pagesizes.A4,
    rightMargin=40,
    leftMargin=40,
    topMargin=40,
    bottomMargin=40
)

elements = []

styles = getSampleStyleSheet()

title_style = ParagraphStyle(
    'TitleStyle',
    parent=styles['Heading1'],
    fontSize=16,
    spaceAfter=12
)

caption_style = ParagraphStyle(
    'CaptionStyle',
    parent=styles['Normal'],
    fontSize=10,
    textColor=colors.grey,
    spaceAfter=18
)

elements.append(Paragraph("Table 1. Pulse Dispersion Engineering Summary", title_style))
elements.append(Paragraph(
    "Comparison of spectral phase types and resulting pulse characteristics.",
    caption_style
))

# ----------------------------------------------------------
# Construct Table
# ----------------------------------------------------------

table = Table(data, repeatRows=1)

table.setStyle(TableStyle([

    # Header
    ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#E6E6E6")),
    ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
    ('FONTSIZE', (0,0), (-1,0), 9),
    ('ALIGN', (0,0), (-1,0), 'CENTER'),

    # Body
    ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
    ('FONTSIZE', (0,1), (-1,-1), 8),
    ('ALIGN', (0,1), (-1,-1), 'CENTER'),

    # Clean journal-style lines
    ('LINEBELOW', (0,0), (-1,0), 1.2, colors.black),
    ('LINEABOVE', (0,-1), (-1,-1), 1.2, colors.black),
    ('INNERGRID', (0,0), (-1,-1), 0.25, colors.lightgrey),

    # Padding
    ('LEFTPADDING', (0,0), (-1,-1), 4),
    ('RIGHTPADDING', (0,0), (-1,-1), 4),
    ('TOPPADDING', (0,0), (-1,-1), 4),
    ('BOTTOMPADDING', (0,0), (-1,-1), 4),
]))

elements.append(table)

doc.build(elements)

print("Publication-ready table saved to:")
print(pdf_path)
